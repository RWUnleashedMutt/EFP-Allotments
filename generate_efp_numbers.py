"""
generate_efp_numbers.py
-----------------------
Reads a Square category-sales CSV export and produces an Excel workbook
with one sheet per brand, matching the format of EFP_Numbers.xlsx.
 
Each sheet contains:
    Location | Brand | Total | % of Total
    ...followed by a blank row, then Coastal and Local summary rows,
    then a budget allocation row showing each group's share of the brand budget.
 
Usage:
    python generate_efp_numbers.py
 
Edit the CONFIG section below to point at your files and adjust brands /
channel mapping / budgets as needed.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIG  –  edit these values as needed
# ─────────────────────────────────────────────

CSV_PATH = "category-sales-2026-05-19-2026-06-19.csv"
OUTPUT_PATH = "EFP_Numbers_All_Brands.xlsx"

# Brands + their allocated budgets
BRANDS = {
    "Natures Logic":   1000,
    "Open Farm":       2000,
    "Small Batch":     1100,
    "Vital Essentials": 2500,
    "Fromm":           1200,
}

CHANNEL_TO_STORE = {
    "City Market: DTR":          "CM",
    "Crabtree Valley Mall":      "CVM",
    "Downtown Durham":           "DTD",
    "Parkway Plaza":             "PP",
    "The Streets at Southpoint": "SS",
    "Southport - Tidewater":     "SP",
    "Front Street":              "MF",
    "Lake Boone":                "LB",
    "Stonehenge Market":         "SH",
    "Landfall Shopping Center":  "LF",
    "Crescent Commons":          "CC",
}

COASTAL_STORES = {"MF", "SP", "LF"}
LOCAL_STORES = {"CM", "CVM", "DTD", "PP", "SS", "LB", "SH", "CC"}

# ─────────────────────────────────────────────
#  STYLES
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
COASTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
LOCAL_FILL = PatternFill("solid", fgColor="DDEBF7")
BUDGET_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow
GROUP_FONT = Font(bold=True, italic=True)
BUDGET_FONT = Font(bold=True, italic=True, color="7F6000")
PCT_FORMAT = "0.00%"
DLR_FORMAT = "#,##0.00"

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────


def styled_row(ws, row, fill, font, col1, col2, total, pct):
    for ci in range(1, 5):
        ws.cell(row, ci).fill = fill
    ws.cell(row, 1, col1).font = font
    ws.cell(row, 2, col2).font = font
    ws.cell(row, 2).fill = fill
    c = ws.cell(row, 3, total)
    c.number_format = DLR_FORMAT
    c.font = font
    c.fill = fill
    p = ws.cell(row, 4, pct)
    p.number_format = PCT_FORMAT
    p.font = font
    p.fill = fill


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    print(f"Reading {CSV_PATH} ...")
    df = pd.read_csv(CSV_PATH)
    df["Gross Sales"] = (
        df["Gross Sales"].astype(str)
        .str.replace(r"[\$,]", "", regex=True)
        .astype(float)
    )

    wb = Workbook()
    wb.remove(wb.active)

    for brand, budget in BRANDS.items():
        brand_df = df[df["Category"] == brand]

        rows = []
        for channel, code in CHANNEL_TO_STORE.items():
            match = brand_df[brand_df["Channel"] == channel]
            total = match["Gross Sales"].sum() if not match.empty else 0.0
            rows.append({"Location": code, "Total": total})
        summary = pd.DataFrame(rows)
        grand_total = summary["Total"].sum()
        summary["pct"] = summary["Total"] / grand_total if grand_total else 0.0

        coastal_total = summary[summary["Location"].isin(
            COASTAL_STORES)]["Total"].sum()
        local_total = summary[summary["Location"].isin(
            LOCAL_STORES)]["Total"].sum()
        coastal_pct = coastal_total / grand_total if grand_total else 0.0
        local_pct = local_total / grand_total if grand_total else 0.0

        coastal_budget = budget * coastal_pct
        local_budget = budget * local_pct

        ws = wb.create_sheet(title=brand[:31])

        # Header
        headers = ["Location", "Brand", "Total", "% of Total"]
        col_widths = [12, 32, 14, 14]
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Data rows
        for ri, row in enumerate(summary.itertuples(index=False), 2):
            ws.cell(ri, 1, row.Location)
            ws.cell(ri, 2, brand)
            ws.cell(ri, 3, row.Total).number_format = DLR_FORMAT
            ws.cell(ri, 4, row.pct).number_format = PCT_FORMAT

        # Grand total row
        tr = len(summary) + 2
        gt = ws.cell(tr, 3, grand_total)
        gt.number_format = DLR_FORMAT
        gt.font = TOTAL_FONT
        pt = ws.cell(tr, 4, 1.0)
        pt.number_format = PCT_FORMAT
        pt.font = TOTAL_FONT

        # Blank separator
        # tr + 1 intentionally empty

        # Coastal summary
        cr = tr + 2
        styled_row(ws, cr, COASTAL_FILL, GROUP_FONT, "Coastal",
                   "MF, SP, LF", coastal_total, coastal_pct)

        # Coastal budget allocation
        cbr = cr + 1
        styled_row(ws, cbr, COASTAL_FILL, BUDGET_FONT,
                   "", f"Coastal allotment of ${budget:,.0f} budget", coastal_budget, coastal_pct)

        # Local summary
        lr = cbr + 2
        styled_row(ws, lr, LOCAL_FILL, GROUP_FONT, "Local",
                   "CM, CVM, DTD, PP, SS, LB, SH, CC", local_total, local_pct)

        # Local budget allocation
        lbr = lr + 1
        styled_row(ws, lbr, LOCAL_FILL, BUDGET_FONT,
                   "", f"Local allotment of ${budget:,.0f} budget", local_budget, local_pct)

        print(
            f"  ✓  {brand:20s}  budget ${budget:,}  →  Coastal ${coastal_budget:,.2f}  |  Local ${local_budget:,.2f}")

    wb.save(OUTPUT_PATH)
    print(f"\nSaved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
